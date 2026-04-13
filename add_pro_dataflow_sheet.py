import shutil
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

warnings.filterwarnings('ignore')

source_file = 'CreditRating_FeatureAnalysis_v6.xlsx'
dest_file = 'CreditRating_FeatureAnalysis_v7.xlsx'

try:
    shutil.copy(source_file, dest_file)
except Exception as e:
    print(f"Error copying file: {e}")

wb = load_workbook(dest_file)

sheet_title = "Data Flow Architecture"
if "Pipeline Data Flow" in wb.sheetnames:
    del wb["Pipeline Data Flow"]
if sheet_title in wb.sheetnames:
    del wb[sheet_title]
ws = wb.create_sheet(title=sheet_title)

# Fonts
FONT_MAIN_TITLE = Font(name='Segoe UI', bold=True, size=16, color='FFFFFF')
FONT_STEP_NUM = Font(name='Segoe UI', bold=True, size=24, color='FFFFFF')
FONT_STEP_TITLE = Font(name='Segoe UI', bold=True, size=13, color='102A43')
FONT_BODY_BOLD = Font(name='Segoe UI', bold=True, size=11, color='334E68')
FONT_BODY = Font(name='Segoe UI', bold=False, size=11, color='486581')
FONT_ARROW = Font(name='Segoe UI', bold=True, size=14, color='9FB3C8')

# Fills
FILL_MAIN = PatternFill(start_color='102A43', end_color='102A43', fill_type='solid') # Deep Navy
FILL_STEP_BG = PatternFill(start_color='243B53', end_color='243B53', fill_type='solid') # Slate Navy
FILL_HEADER = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid') # Very Light Blue/Grey
FILL_CONTENT = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid') # White
FILL_ARROW_BG = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Borders
THIN_PALE = Side(style='thin', color="D9E2EC")
BORDER_BOX = Border(left=THIN_PALE, right=THIN_PALE, top=THIN_PALE, bottom=THIN_PALE)
BORDER_LEFT_ONLY = Border(left=Side(style='medium', color="243B53"))

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
ALIGN_LEFT_NO_INDENT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Define data
flow_data = [
    {
        "num": "01",
        "title": "THU THẬP & TỔ CHỨC HỒ SƠ TÍN DỤNG (INPUT DATA)",
        "items": [
            "• [Hồ sơ Báo cáo Tài Chính] Tập hợp các dữ liệu nền tảng về Lợi nhuận, Đòn bẩy, Thanh khoản.",
            "• [Năng lực Quản trị Số liệu] Kiểm thử xuất xứ dữ liệu minh bạch thông qua hệ thống API tự động."
        ]
    },
    {
        "num": "02",
        "title": "THẨM ĐỊNH LÀM SẠCH BÁO CÁO & BENCHMARKING",
        "items": [
            "• [Khử lệch độ nhạy Ngành] Đặt báo cáo của doanh nghiệp vào khung tham chiếu chung của ngành nhằm tạo mặt bằng rủi ro công bằng (Sector-relative Benchmarking).",
            "• [Kiểm thử sức chịu đựng] Đánh giá rủi ro ngầm theo giả định khủng hoảng thanh khoản để test giới hạn chịu đựng (Stress Testing)."
        ]
    },
    {
        "num": "03",
        "title": "ÁNH XẠ PHÂN LỚP RỦI RO (MASTER SCALE MAP)",
        "items": [
            "🟢 [Nhóm An Toàn - IG] Cấp độ được bảo lãnh phân phối vào các danh mục quỹ hưu trí chi phí thấp và an toàn tuyệt đối.",
            "🟡 [Nhóm Đầu Cơ - HY] Nhóm lợi suất rủi ro cao, bù đắp bằng lãi lớn nhưng dễ nứt gãy cấu trúc khi vĩ mô xấu đi.",
            "🔴 [Nhóm Vỡ Nợ - D] Điểm giới hạn sinh tồn, mốc đánh giá khắt khe nhất để cảnh báo mất thanh khoản."
        ]
    },
    {
        "num": "04",
        "title": "THIẾT LẬP QUỸ ĐẠO CẤP HẠNG (TRAJECTORY SETUP)",
        "items": [
            "• [Động lượng rủi ro thời gian] Phân tích xu hướng phục hồi qua chuỗi lịch sử quá khứ thay vì dựa vào lát cắt tĩnh của báo cáo tài chính quý gần nhất."
        ]
    },
    {
        "num": "05",
        "title": "HỆ THỐNG CỐT LÕI ĐÁNH GIÁ TÍN NHIỆM",
        "items": [
            "• [Phân tích Mô phỏng Chu kỳ] Giám sát năng lực tài chính dựa trên khả năng trả nợ và dòng tiền tương lai.",
            "• [Kiểm soát Covenants Rủi ro] Thiết lập màng lọc chặn tự động những phán đoán trái ngược với diễn biến dòng tiền thực tế."
        ]
    },
    {
        "num": "06",
        "title": "QUYẾT ĐỊNH CẤP VỐN & GIẢI TRÌNH (CREDIT MEMO)",
        "items": [
            "• [Xác suất Phân bổ] Xuất khuyến nghị phân bổ hạn mức vốn lý tưởng đối với các khoản tín dụng lớn.",
            "• [Tờ trình Rủi ro - Credit Memo] Bóc tách bằng chứng định lượng lý giải một cách sắc bén tại sao doanh nghiệp bị hạ/nâng điểm tín nhiệm."
        ]
    }
]

# Write Main Title
ws.merge_cells("B2:G3")
title_cell = ws.cell(row=2, column=2, value="HỆ THỐNG XẾP HẠNG TÍN NHIỆM DOANH NGHIỆP\nTRƯỜNG DỮ LIỆU ĐÁNH GIÁ (CORPORATE CREDIT RATING PIPELINE)")
title_cell.fill = FILL_MAIN
title_cell.font = FONT_MAIN_TITLE
title_cell.alignment = ALIGN_CENTER

curr_row = 5

for i, block in enumerate(flow_data):
    num_items = len(block["items"])
    total_rows = num_items + 1 
    
    end_row = curr_row + total_rows - 1
    
    # Write Step Number
    ws.merge_cells(start_row=curr_row, start_column=2, end_row=end_row, end_column=2)
    num_cell = ws.cell(row=curr_row, column=2, value=block["num"])
    num_cell.fill = FILL_STEP_BG
    num_cell.font = FONT_STEP_NUM
    num_cell.alignment = ALIGN_CENTER
    
    # Write Title
    ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=7)
    th_cell = ws.cell(row=curr_row, column=3, value="  " + block["title"])
    th_cell.fill = FILL_HEADER
    th_cell.font = FONT_STEP_TITLE
    th_cell.alignment = ALIGN_LEFT_NO_INDENT
    for col in range(3, 8):
        c = ws.cell(row=curr_row, column=col)
        c.border = Border(left=Side(style='medium', color="243B53") if col==3 else None,
                          top=THIN_PALE, right=THIN_PALE if col==7 else None)
    ws.row_dimensions[curr_row].height = 25
    
    # Write Items
    item_row = curr_row + 1
    for item in block["items"]:
        ws.merge_cells(start_row=item_row, start_column=3, end_row=item_row, end_column=7)
        it_cell = ws.cell(row=item_row, column=3, value=item)
        it_cell.fill = FILL_CONTENT
        if "🟢" in item or "🟡" in item or "🔴" in item:
            it_cell.font = FONT_BODY_BOLD
        else:
            it_cell.font = FONT_BODY
        it_cell.alignment = ALIGN_LEFT
        for col in range(4, 8):
            c = ws.cell(row=item_row, column=col)
            c.border = Border(right=THIN_PALE if col==7 else None, bottom=THIN_PALE)
            
        it_cell.border = Border(left=Side(style='medium', color="243B53"), bottom=THIN_PALE)
            
        ws.row_dimensions[item_row].height = 30
        item_row += 1
        
    curr_row = end_row + 1
    
    # Draw Arrow if not last
    if i < len(flow_data) - 1:
        ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=7)
        arr_cell = ws.cell(row=curr_row, column=3, value="▼")
        arr_cell.font = FONT_ARROW
        arr_cell.alignment = ALIGN_CENTER
        ws.row_dimensions[curr_row].height = 20
        curr_row += 1
        
# Set column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 22

ws.sheet_view.showGridLines = False

# Give tab a vibrant color
ws.sheet_properties.tabColor = "102A43"

wb.save(dest_file)
print(f"Professional design saved to {dest_file}")
