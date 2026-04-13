import shutil
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

warnings.filterwarnings('ignore')

source_file = 'CreditRating_FeatureAnalysis_v5.xlsx'
dest_file = 'CreditRating_FeatureAnalysis_v6.xlsx'

# Copy the file to v6 before modifying
try:
    shutil.copy(source_file, dest_file)
except Exception as e:
    print(f"Error copying file: {e}")

# Load the workbook
wb = load_workbook(dest_file)

# We will recreate the sheet for the Generalized Flow
sheet_title = "Pipeline Data Flow"
if sheet_title in wb.sheetnames:
    del wb[sheet_title]
ws = wb.create_sheet(title=sheet_title)

# Prepare Styles
FONT_TITLE = Font(name='Arial', bold=True, size=16, color='FFFFFF')
FONT_HEADER = Font(name='Arial', bold=True, size=13, color='FFFFFF')
FONT_BODY_BOLD = Font(name='Arial', bold=True, size=11, color='333333')
FONT_BODY = Font(name='Arial', bold=False, size=11, color='333333')

FILL_TITLE = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')      
FILL_INPUT = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')      
FILL_PREPROCESS = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') 
FILL_MAPPING = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') 
FILL_MODEL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')      
FILL_OUTPUT = PatternFill(start_color='E6B8B7', end_color='E6B8B7', fill_type='solid')     
FILL_ARROW = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')     

ALIGN_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)

THICK_BORDER_SIDE = Side(style='medium', color="000000")
BORDER_BLOCK = Border(left=THICK_BORDER_SIDE, right=THICK_BORDER_SIDE, top=THICK_BORDER_SIDE, bottom=THICK_BORDER_SIDE)

# Content layout: Strictly Business/General, No Technical Jargon
flow_data = [
    # Level 1: TITLE
    ("HỆ THỐNG XẾP HẠNG TÍN NHIỆM DOANH NGHIỆP (CORPORATE CREDIT RATING PIPELINE)", FILL_TITLE, FONT_TITLE, 3),
    
    # ------------------
    # BLOCK 1: INPUT
    # ------------------
    ("TẦNG 1: THU THẬP & TỔ CHỨC HỒ SƠ TÍN DỤNG", FILL_INPUT, FONT_HEADER, 2),
    ("[Hồ sơ Báo cáo Tài Chính] Tập hợp các dữ liệu nền tảng về Lợi nhuận, Đòn bẩy, Thanh khoản.", FILL_INPUT, FONT_BODY_BOLD, 2),
    ("[Năng lực Quản trị Số liệu] Kiểm thử xuất xứ dữ liệu minh bạch thông qua hệ thống tích hợp thông tin tự động.", FILL_INPUT, FONT_BODY, 2),
    
    ("⬇", FILL_ARROW, FONT_TITLE, 2),
    
    # ------------------
    # BLOCK 2: PREPROCESS
    # ------------------
    ("TẦNG 2: THẨM ĐỊNH LÀM SẠCH BÁO CÁO & BENCHMARKING", FILL_PREPROCESS, FONT_HEADER, 2),
    ("[Khử sai lệch giữa các nhóm Ngành] Đặt báo cáo của doanh nghiệp vào khung tham chiếu chung của ngành (VD: Tài chính vs Sản xuất).", FILL_PREPROCESS, FONT_BODY_BOLD, 2),
    ("[Kiểm thử sức chịu đựng - Stress Testing] Đánh giá rủi ro ngầm khi đối mặt rủi ro hiếm có trên thị trường.", FILL_PREPROCESS, FONT_BODY, 2),
    
    ("⬇", FILL_ARROW, FONT_TITLE, 2),
    
    # ------------------
    # BLOCK 3: TARGET MAPPING 
    # ------------------
    ("TẦNG 3: ÁNH XẠ NHÓM RỦI RO (MASTER SCALE - 3 GROUPS)", FILL_MAPPING, FONT_HEADER, 2),
    ("[Phân lớp An toàn: Đầu tư (IG)] Trọng tâm dành cho huy động vốn chi phí thấp, an toàn tuyệt đối.", FILL_MAPPING, FONT_BODY_BOLD, 2),
    ("[Phân lớp Rủi ro: Đầu cơ (HY)] Đòi hỏi quản trị kỹ rủi ro rạn nứt cấu trúc vốn để được bồi thường bằng lãi suất cao.", FILL_MAPPING, FONT_BODY_BOLD, 2),
    ("[Phân lớp Báo động: Vỡ nợ (D)] Theo sát các hệ liệm cảnh báo sớm khi năng lực thanh toán vi phạm giới hạn sinh tồn.", FILL_MAPPING, FONT_BODY_BOLD, 2),
    
    ("⬇", FILL_ARROW, FONT_TITLE, 2),
    
    # ------------------
    # BLOCK 4: SEQUENCING
    # ------------------
    ("TẦNG 4: THIẾT LẬP QUỸ ĐẠO CẤP HẠNG (CREDIT TRAJECTORY)", FILL_PREPROCESS, FONT_HEADER, 2),
    ("[Khung thời gian Động lượng] Thay vì chấm điểm một khoảnh khắc tĩnh, gom chuỗi lịch sử thăng hạng / tụt hạng của doanh nghiệp để vẽ bức tranh xu hướng rủi ro.", FILL_PREPROCESS, FONT_BODY_BOLD, 2.5),

    ("⬇", FILL_ARROW, FONT_TITLE, 2),

    # ------------------
    # BLOCK 5: MODELING
    # ------------------
    ("TẦNG 5: HỆ THỐNG CỐT LÕI ĐÁNH GIÁ CHẤT LƯỢNG", FILL_MODEL, FONT_HEADER, 2),
    ("[Mô phỏng Phân tích Chu kỳ] Theo sát khả năng hấp thụ rủi ro dựa vào chu kỳ kinh tế và khả năng trả nợ tương lai.", FILL_MODEL, FONT_BODY_BOLD, 2),
    ("[Quy tắc Quản trị Rủi ro Tín dụng] Giăng thiết bị cảnh báo tự động thông qua các 'Covenants' (Chốt chặn vi phạm) để bảo vệ quyết định, chống chấm điểm sai lệch với thực tế.", FILL_MODEL, FONT_BODY, 2),
    
    ("⬇", FILL_ARROW, FONT_TITLE, 2),
    
    # ------------------
    # BLOCK 6: OUTPUT
    # ------------------
    ("TẦNG 6: QUYẾT ĐỊNH CẤP VỐN & GIẢI TRÌNH BÁO CÁO (CREDIT MEMO)", FILL_OUTPUT, FONT_HEADER, 2),
    ("[Giám định Xác suất] Xuất khuyến nghị phân phối Vốn của một khoản vay dựa vào điểm xếp hạng Rủi ro mới.", FILL_OUTPUT, FONT_BODY_BOLD, 2),
    ("[Tờ trình Minh bạch Yếu tố Hạ hạng] Nêu rõ các bằng chứng định lượng lý giải một quyết định nâng hoặc hạ tín nhiệm lên Cấp phê duyệt.", FILL_OUTPUT, FONT_BODY, 2),
]

current_row = 2

def write_block(ws, text, fill, font, height_multiplier, row_idx):
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=7)
    cell = ws.cell(row=row_idx, column=2)
    cell.value = text
    cell.fill = fill
    cell.font = font
    cell.alignment = ALIGN_CENTER_WRAP 
    
    # Apply border
    for col in range(2, 8):
        current = ws.cell(row=row_idx, column=col)
        current.border = BORDER_BLOCK
        
    ws.row_dimensions[row_idx].height = 15 * height_multiplier

for text, fill, font, h_mul in flow_data:
    write_block(ws, text, fill, font, h_mul, current_row)
    current_row += 1

# Formatting columns
ws.column_dimensions['A'].width = 5
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws.column_dimensions[col].width = 20

ws.sheet_view.showGridLines = False

wb.save(dest_file)
print(f"General Data Flow Diagram updated in {dest_file}")
