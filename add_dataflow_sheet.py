import shutil
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

warnings.filterwarnings('ignore')

source_file = 'CreditRating_FeatureAnalysis_v4.xlsx'
dest_file = 'CreditRating_FeatureAnalysis_v5.xlsx'

# Copy the file to v5 before modifying
try:
    shutil.copy(source_file, dest_file)
except Exception as e:
    print(f"Error copying file: {e}")

# Load the workbook
wb = load_workbook(dest_file)

# We will recreate the sheet to include the Grouping block
sheet_title = "Pipeline Data Flow"
if sheet_title in wb.sheetnames:
    del wb[sheet_title]
ws = wb.create_sheet(title=sheet_title)

# Prepare Styles
FONT_TITLE = Font(name='Arial', bold=True, size=16, color='FFFFFF')
FONT_HEADER = Font(name='Arial', bold=True, size=12, color='FFFFFF')
FONT_BODY_BOLD = Font(name='Arial', bold=True, size=11, color='333333')
FONT_BODY = Font(name='Arial', bold=False, size=11, color='333333')

FILL_TITLE = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')      
FILL_INPUT = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')      
FILL_PREPROCESS = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') 
FILL_MAPPING = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Light Peach for grouping mapping
FILL_MODEL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')      
FILL_OUTPUT = PatternFill(start_color='E6B8B7', end_color='E6B8B7', fill_type='solid')     
FILL_ARROW = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')     

ALIGN_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)

THICK_BORDER_SIDE = Side(style='medium', color="000000")
BORDER_BLOCK = Border(left=THICK_BORDER_SIDE, right=THICK_BORDER_SIDE, top=THICK_BORDER_SIDE, bottom=THICK_BORDER_SIDE)

# Content layout (Row by Row)
flow_data = [
    # Level 1: TITLE
    ("HỆ THỐNG MÔ HÌNH HÓA ĐÁNH GIÁ TÍN NHIỆM (TRANSFORMER-BiLSTM-FUZZY)", FILL_TITLE, FONT_TITLE, 3),
    
    # ------------------
    # BLOCK 1: INPUT
    # ------------------
    ("TẦNG 1: DỮ LIỆU ĐẦU VÀO (INPUT LAYER)", FILL_INPUT, FONT_HEADER, 2),
    ("[Nguồn Raw Data] Báo cáo tài chính, Dữ liệu ngành, Dữ liệu thị trường vĩ mô", FILL_INPUT, FONT_BODY_BOLD, 2),
    ("[Scrapers] FiinRatings Scraper / Các cổng API Dữ liệu", FILL_INPUT, FONT_BODY, 2),
    
    ("⬇", FILL_ARROW, FONT_TITLE, 2),
    
    # ------------------
    # BLOCK 2: PREPROCESS
    # ------------------
    ("TẦNG 2: TIỀN XỬ LÝ & CHUẨN HÓA (PREPROCESSING LAYER)", FILL_PREPROCESS, FONT_HEADER, 2),
    ("[Chuẩn hóa Ngành - Z-score] Sector-relative Normalization (Khử sai lệch giữa công nghiệp nặng và dịch vụ)", FILL_PREPROCESS, FONT_BODY_BOLD, 2),
    ("[Feature Engineering] Khởi tạo Temporal Deltas và Quản trị nhiễu Outlier", FILL_PREPROCESS, FONT_BODY, 2),
    
    ("⬇", FILL_ARROW, FONT_TITLE, 2),
    
    # ------------------
    # BLOCK 3: TARGET MAPPING (NEW BLOCK FOR 3-GROUPS)
    # ------------------
    ("TẦNG 3: ÁNH XẠ TARGET & PHÂN NHÓM RỦI RO (MASTER SCALE 3-GROUPS)", FILL_MAPPING, FONT_HEADER, 2),
    ("[Nhóm 1: INVESTMENT GRADE - IG] Cấp độ An toàn (Thường từ AAA đến BBB-). Đặc trưng: Các quỹ phòng hộ/hưu trí lớn cấu trúc tỷ trọng mua tỷ lệ cao.", FILL_MAPPING, FONT_BODY_BOLD, 2.5),
    ("[Nhóm 2: HIGH YIELD - HY] Nhóm Đầu cơ / Nhạy cảm (Thường từ BB+ đến C). Đặc trưng: Lợi tức cao (Junk) nhưng biên độ rạn nứt cấu trúc vốn mỏng manh khi vĩ mô đảo chiều.", FILL_MAPPING, FONT_BODY_BOLD, 2.5),
    ("[Nhóm 3: DEFAULT / DISTRESSED - D] Vỡ nợ / Suy kiệt trầm trọng. Đặc trưng: Mất khả năng thanh toán cấu trúc vốn, bắt đầu đưa vào xử lý thanh lý/tái cơ cấu.", FILL_MAPPING, FONT_BODY_BOLD, 2.5),
    ("[Data Augmentation] Bắt buộc dùng TimeGAN/SMOTE bơm sinh dữ liệu giả lập cho nhóm thiểu số nhằm cân bằng sự học của thuật toán.", FILL_MAPPING, FONT_BODY, 2),
    
    ("⬇", FILL_ARROW, FONT_TITLE, 2),
    
    # ------------------
    # BLOCK 4: SEQUENCING
    # ------------------
    ("TẦNG 4: CHUỖI HÓA KHÔNG GIAN THỜI GIAN (4D WINDOWING)", FILL_PREPROCESS, FONT_HEADER, 2),
    ("[Sliding Window] Nhóm chuỗi thời báo cáo (Company - Ticker - Time) theo shape dạng (Batch, Timesteps, Features) để theo dõi động lượng sức khỏe công ty (Credit Trajectory)", FILL_PREPROCESS, FONT_BODY_BOLD, 2),

    ("⬇", FILL_ARROW, FONT_TITLE, 2),

    # ------------------
    # BLOCK 5: MODELING
    # ------------------
    ("TẦNG 5: THUẬT TOÁN HỌC SÂU (DEEP LEARNING ENGINE)", FILL_MODEL, FONT_HEADER, 2),
    ("[RoPE Transformer] Mạng chú ý encode cấu trúc tuần tự các báo cáo", FILL_MODEL, FONT_BODY, 2),
    ("[Bi-LSTM] Khai thác bộ nhớ dài hạn 2 chiều trước/sau của chu kỳ kinh tế", FILL_MODEL, FONT_BODY, 2),
    ("[Fuzzy Logic Layer] Đưa thêm luật mờ (Fuzzy laws/Covenants) để đánh sập Persistence Bias (Chỉ bám vào rank cũ)", FILL_MODEL, FONT_BODY_BOLD, 2),
    ("[CORN Loss] Phạt ngặt nghèo các dự báo trượt nhóm. Vd: Dự trữ nhầm HY thành IG bị phạt cấp số nhân so với sai số cơ bản.", FILL_MODEL, FONT_BODY_BOLD, 2),
    
    ("⬇", FILL_ARROW, FONT_TITLE, 2),
    
    # ------------------
    # BLOCK 6: OUTPUT
    # ------------------
    ("TẦNG 6: KẾT QUẢ ĐẦU RA (OUTPUT & CREDIT MEMO)", FILL_OUTPUT, FONT_HEADER, 2),
    ("[Xác suất Ordinal 3 Nhóm] Mapping kết quả đầu ra thành Xác Suất Phục vụ Cho vay/Đầu tư", FILL_OUTPUT, FONT_BODY_BOLD, 2),
    ("[Tờ trình Giải thích SHAP] Khai xuất Đóng góp ngược của từng Feature làm cứ liệu giải trình rủi ro (Risk Distillation)", FILL_OUTPUT, FONT_BODY, 2),
]

current_row = 2

def write_block(ws, text, fill, font, height_multiplier, row_idx):
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=7)
    cell = ws.cell(row=row_idx, column=2)
    cell.value = text
    cell.fill = fill
    cell.font = font
    cell.alignment = ALIGN_CENTER_WRAP 
    
    # Apply border
    for col in range(2, 8):
        current = ws.cell(row=row_idx, column=col)
        current.border = BORDER_BLOCK
        
    ws.row_dimensions[row_idx].height = 15 * height_multiplier

for text, fill, font, h_mul in flow_data:
    write_block(ws, text, fill, font, h_mul, current_row)
    current_row += 1

# Formatting columns
ws.column_dimensions['A'].width = 5
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws.column_dimensions[col].width = 20

ws.sheet_view.showGridLines = False

wb.save(dest_file)
print(f"Data Flow Diagram updated in {dest_file}")
