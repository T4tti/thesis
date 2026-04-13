import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.impute import SimpleImputer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

# 1. Load data
df = pd.read_csv('data/processed/merged_credit_rating_common_3groups.csv')

# 2. Identify features and target
target = 'rating_detail'
features = [
    'current_ratio', 'debt_equity_ratio', 'gross_profit_margin', 
    'operating_profit_margin', 'ebit_margin', 'pretax_profit_margin', 
    'net_profit_margin', 'asset_turnover', 'roe', 'roa', 
    'operating_cashflow_ps', 'free_cashflow_ps'
]

X = df[features]
y = df[target]

# 3. Preprocess
imputer = SimpleImputer(strategy='median')
X_imputed = imputer.fit_transform(X)
X_imputed = pd.DataFrame(X_imputed, columns=features)

le = LabelEncoder()
y_encoded = le.fit_transform(y)

# 4. Train Model
model = RandomForestClassifier(n_estimators=100, random_state=42, max_depth=10)
model.fit(X_imputed, y_encoded)

# 5. Extract Feature Importances (Gini importance)
importances = model.feature_importances_

# Create a dataframe for features
feat_df = pd.DataFrame({
    'Trường': features,
    'Importance': importances
})
feat_df = feat_df.sort_values(by='Importance', ascending=False).reset_index(drop=True)
feat_df['Rank'] = feat_df.index + 1

# Descriptions
desc = {
    'current_ratio': 'Năng lực thanh khoản ngắn hạn (Short-term Liquidity). Đánh giá "tấm đệm" lưu chuyển tiền tệ (cash flow buffer) để chống chịu các cú sốc thanh khoản trong 12 tháng tới. Tỷ lệ < 1.0 (hoặc thấp so với trung bình ngành) báo hiệu rủi ro tái cấp vốn (refinancing risk) khẩn cấp.',
    'debt_equity_ratio': 'Rủi ro cơ cấu vốn & Đòn bẩy tài chính (Financial Leverage). Là một trong những "hard-constraint" (chốt chặn) quan trọng nhất. Đòn bẩy cao phản ánh lớp vốn đệm (equity cushion) mỏng, làm tăng theo cấp số nhân Xác suất vỡ nợ (Probability of Default - PD) khi vĩ mô suy thoái.',
    'gross_profit_margin': 'Lợi thế cạnh tranh & Gia tăng giá trị (Pricing Power). Là chỉ báo cốt lõi của khả năng chuyển dịch chi phí đầu vào lên người tiêu dùng. Biên gộp càng ổn định qua các chu kỳ kinh tế chứng tỏ rào cản gia nhập ngành lớn, giúp cải thiện Điểm rủi ro kinh doanh (Business Risk Profile).',
    'operating_profit_margin': 'Hiệu quả vận hành (Operational Efficiency). Trong thẻ điểm tín nhiệm (Credit Scorecard), biên hoạt động đánh giá sức khỏe sinh lời thuần túy từ mảng kinh doanh lõi, loại bỏ các yếu tố thu nhập bất thường. Sự ổn định của chỉ số này ảnh hưởng mạnh tới xếp loại Ổn định (Outlook Stable).',
    'ebit_margin': 'Khả năng sinh lời lõi phục vụ trả nợ (Debt Servicing Base). EBIT là nguồn thu vô cùng quan trọng dùng để tính EBIT Interest Coverage (Khả năng thanh toán lãi vay). Biên EBIT càng dày chứng tỏ bộ đệm sẵn sàng thanh toán nghĩa vụ nợ cố định (Fixed-charge cover) càng vững mạnh.',
    'pretax_profit_margin': 'Mức sinh lời trước mức thuế phải nộp. Hữu ích cho nghiệp vụ so sánh tính điểm tín nhiệm ngang hàng (Peer comparison) trong các ngành có lá chắn thuế (Tax shield) biến động mạnh giữa các quốc gia hoặc chu kỳ kinh tế.',
    'net_profit_margin': 'Hiệu quả tạo vốn tự có (Capital Generation Capacity). Biên ròng quyết định dòng lợi nhuận giữ lại (Retained Earnings) dùng để bổ sung trực tiếp vào Vốn chủ sở hữu, từ đó cấu trúc lại sức khỏe của bảng cân đối kế toán. Giảm sút sẽ cảnh báo cạn kiệt vốn (Capital Depletion).',
    'asset_turnover': 'Hiệu suất sử dụng vốn (Asset Efficiency/Capital Intensity). Đặc biệt cực kỳ quan trọng với công nghiệp nặng/bán lẻ. Vòng quay sụt giảm là "Red Flag" về hàng tồn kho/khoản phải thu ứ đọng, báo trước dòng tiền HĐKD sẽ suy yếu.',
    'roe': 'Lợi nhuận cổ đông (Return on Equity). Trong nghiệp vụ Credit Rating, ROE quá cao có thể là hệ quả của việc lạm dụng nợ vay (hiệu ứng DuPont). Cơ quan xếp hạng phân tích chéo ROE với Đòn bẩy để nhận diện các thủ thuật tối ưu hóa chỉ số che bảo rủi ro vỡ nợ.',
    'roa': 'Sinh lời tổng thể (Return on Assets). Đánh giá chân thực năng lực quản trị vốn (Management Capability) trên mọi đồng tài trợ (Dù từ Nợ hay Vốn) mà không bị che lấp / bóp méo bởi hiệu ứng đòn bẩy tài chính.',
    'operating_cashflow_ps': 'Sức khỏe thực chất. Dòng tiền HĐKD (CFO) là chỉ báo Sinh tồn (Cash is King). CFO có thể phát hiện thủ thuật kế toán bóp méo lợi nhuận ghi sổ (Earnings Management). Dòng tiền CFO âm kéo dài trực tiếp dẫn đến rủi ro vỡ nợ kỹ thuật (Technical Default) dù có lãi.',
    'free_cashflow_ps': 'Tự chủ tài chính (Stand-alone Credit Profile). FCF (Dòng tiền tự do) đo lường nguồn thặng dư sau chi phí đầu tư duy trì (CapEx). FCF dồi dào chứng tỏ khả năng tự trả nền gốc nợ mà không phải đảo nợ (Roll-over). FCF âm thường là đặc trưng của khu vực doanh nghiệp trái phiếu rác (Junk / Speculative Grade).'
}

# 6. Generate DataFrame
out_data = []

# Title metadata
meta_title = 'TRƯỜNG DỮ LIỆU – Chuẩn nghiệp vụ xếp hạng doanh nghiệp (S&P/Moody\'s/FiinRatings)'

# Metadata columns
metadata_cols = [
    ('TARGET: rating_detail', 'String (Categorical)', 'Nhãn mục tiêu: Thuộc các cấp độ Non-Investment Grade (HY), Investment Grade (IG) theo khung Master Scale nghiệp vụ chuẩn.', 'Biến học tập lõi cho hàm CORN (Conditional Ordinal Regression)'),
    ('META: company_name', 'String', 'Tên tổ chức phát hành (Issuer). Định danh thực thể độc lập phân tích.', 'Tránh rỉ rỉ thông tin trong chia tập Train/Test theo chuỗi'),
    ('META: ticker', 'String', 'Mã chứng khoán / Mã tham chiếu', 'Định danh chuỗi thời gian của entity'),
    ('META: rating_agency', 'String', 'Cơ quan cấp hạng. Yếu tố chuẩn hóa chênh lệch giữa các phương pháp của từng Agency.', 'Condition Feature/Embedding cho mô hình'),
    ('META: rating_date', 'Timeline Feature', 'Ngày cấp hạng. Mốc thời gian kích hoạt cảnh báo.', 'Trục X (Temporal Axis) cho mô hình BiLSTM, chuỗi sliding window'),
    ('META: sector', 'String (Categorical)', 'Nhóm ngành hoạt động (Business Sector). Xử lý rủi ro ngành (Industry Risk) trong khung rủi ro.', 'Phân tán độ tương đương rủi ro đòn bẩy giữa các Ngành'),
    ('META: source', 'String', 'Nguồn thu thập dữ liệu', 'Hỗ trợ audit xuất xứ số liệu')
]
for col, dtype, desc_t, example in metadata_cols:
    out_data.append({
        meta_title: col,
        'Kiểu dữ liệu': dtype,
        'Mô tả phân tích (Chuẩn nghiệp vụ tín nhiệm)': desc_t,
        'Góc nhìn Rủi ro (Risk Impact)': 'Dữ liệu điều hướng / Meta',
        'Đóng góp (%) / Trọng số (Gini)': 'N/A',
        'Phân lớp Trọng yếu': 'Metadata Target',
        'Insights (Dành cho Transformer-BiLSTM)': example
    })

# Feature columns
for i, row in feat_df.iterrows():
    f = f"FEATURE: {row['Trường']}"
    imp = row['Importance']
    rank = row['Rank']
    
    impact = '...'
    if 'ratio' in f or f in ['FEATURE: roe', 'FEATURE: roa', 'FEATURE: operating_cashflow_ps', 'FEATURE: free_cashflow_ps'] or 'margin' in f:
        impact = 'Năng lực tài chính MẠNH -> Tương quan dương với mức đánh giá Đầu tư (IG). Giảm mạnh Nguy cơ vỡ nợ gốc/lãi.'
    if 'debt' in f:
        impact = 'Gia tăng Áp lực Đòn bẩy. Nếu bung rộng báo hiệu cấu trúc vốn mất cân đối nghiêm trọng, kích hoạt tín hiệu hạ bậc.'
        
    out_data.append({
        meta_title: f,
        'Kiểu dữ liệu': 'Financial Ratio',
        'Mô tả phân tích (Chuẩn nghiệp vụ tín nhiệm)': desc.get(row['Trường'], ''),
        'Góc nhìn Rủi ro (Risk Impact)': impact,
        'Đóng góp (%) / Trọng số (Gini)': f"{imp*100:.1f}%",
        'Phân lớp Trọng yếu': f"Rank {rank} / {len(features)}",
        'Insights (Dành cho Transformer-BiLSTM)': f'{"Top Critical (Chốt chặn): LSTM cần chiết xuất ngay các bất thường số liệu (Sector-outliers)." if rank <=3 else ("High Importance: Căn cứ thiết lập nhánh Cross-Feature." if rank <=7 else "Supporting Metric: Tinh chỉnh xác suất ở biên chuẩn.")}'
    })

res_df = pd.DataFrame(out_data)

excel_file = 'CreditRating_FeatureAnalysis_v3.xlsx'
# We have to write the data and explicitly set column widths, wrap text, and background colors.
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    res_df.to_excel(writer, index=False, sheet_name='Feature Analysis')
    
# Now using openpyxl directly to style it
wb = load_workbook(excel_file)
ws = wb['Feature Analysis']

# Colors
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Dark Blue
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)

META_FILL = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid') # Light Gray
FEATURE_GREEN = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid') # Light Green for Top Features
FEATURE_YELLOW = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid') # Light Yellow for Mid
FEATURE_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 1. Format headers (Row 1)
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

# 2. Format Data Rows
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
    is_meta = 'META:' in str(row[0].value) or 'TARGET:' in str(row[0].value)
    
    # Determine the fill for the row
    if is_meta:
        row_fill = META_FILL
    else:
        rank_str = str(row[5].value) # 'Phân lớp Trọng yếu' is col 6
        if 'Rank 1 /' in rank_str or 'Rank 2 /' in rank_str or 'Rank 3 /' in rank_str:
            row_fill = FEATURE_GREEN
        elif 'Rank 4 /' in rank_str or 'Rank 5 /' in rank_str or 'Rank 6 /' in rank_str or 'Rank 7 /' in rank_str:
            row_fill = FEATURE_YELLOW
        else:
            row_fill = FEATURE_WHITE
            
    for cell_idx, cell in enumerate(row):
        cell.fill = row_fill
        cell.border = THIN_BORDER
        if cell_idx in [0, 2, 3, 6]:
            cell.alignment = ALIGN_LEFT
        else:
            cell.alignment = ALIGN_CENTER

# 3. Adjust Column Widths
column_widths = {
    'A': 35, # Tên trường
    'B': 20, # Kiểu
    'C': 50, # Mô tả
    'D': 35, # Góc nhìn
    'E': 15, # Gini
    'F': 15, # Ranking
    'G': 40  # Insights
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width
    
# Freeze panes
ws.freeze_panes = 'A2'

wb.save(excel_file)
print(f"Formatted and saved to {excel_file}")
