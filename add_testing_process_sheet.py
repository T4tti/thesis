import shutil
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

warnings.filterwarnings('ignore')

source_file = 'CreditRating_FeatureAnalysis_v7.xlsx'
dest_file = 'CreditRating_FeatureAnalysis_v8.xlsx'

try:
    shutil.copy(source_file, dest_file)
except Exception as e:
    print(f"Error copying file: {e}")

wb = load_workbook(dest_file)

sheet_title = "Testing Process Protocol"
if sheet_title in wb.sheetnames:
    del wb[sheet_title]
ws = wb.create_sheet(title=sheet_title)

# Fonts
FONT_MAIN_TITLE = Font(name='Segoe UI', bold=True, size=16, color='FFFFFF')
FONT_STEP_NUM = Font(name='Segoe UI', bold=True, size=24, color='FFFFFF')
FONT_STEP_TITLE = Font(name='Segoe UI', bold=True, size=13, color='5B2C6F') # Deep Purple hue for Testing
FONT_BODY_BOLD = Font(name='Segoe UI', bold=True, size=11, color='4A235A')
FONT_BODY = Font(name='Segoe UI', bold=False, size=11, color='486581')
FONT_ARROW = Font(name='Segoe UI', bold=True, size=14, color='B2BABB')

# Fills (Shifting to a sophisticated Plum/Slate theme for Testing, differentiating from Dataflow)
FILL_MAIN = PatternFill(start_color='4A235A', end_color='4A235A', fill_type='solid') # Deep Indigo/Plum
FILL_STEP_BG = PatternFill(start_color='6C3483', end_color='6C3483', fill_type='solid') # Purple
FILL_HEADER = PatternFill(start_color='F4ECF7', end_color='F4ECF7', fill_type='solid') # Pale Purple
FILL_CONTENT = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_ARROW_BG = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Borders
THIN_PALE = Side(style='thin', color="EBDEF0")
BORDER_BOX = Border(left=THIN_PALE, right=THIN_PALE, top=THIN_PALE, bottom=THIN_PALE)
BORDER_LEFT_ONLY = Border(left=Side(style='medium', color="6C3483"))

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
ALIGN_LEFT_NO_INDENT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Define data specific to Model Validation & Testing
flow_data = [
    {
        "num": "01",
        "title": "KIỂM TOÁN TÍNH TOÀN VẸN (DATA SANITY & LEAKAGE CHECK)",
        "items": [
            "• [Ngăn chặn Look-ahead Bias] Rà soát cửa sổ trượt (Sliding Window), tuyệt đối không để dữ liệu tương lai rò rỉ (leakage) ngược vào chuỗi huấn luyện.",
            "• [Cách ly Thực thể Định danh] Đảm bảo dữ liệu báo cáo của 1 Doanh nghiệp (Company_Id) không bị xé nửa vừa rơi vào tập Train, vừa nằm ở tập Test trong cùng một mốc thời gian."
        ]
    },
    {
        "num": "02",
        "title": "PHÂN CHIA TẬP KIỂM THỬ THỜI GIAN THỰC (WALK-FORWARD SPLIT)",
        "items": [
            "• [Phân đoạn Lịch sử - Chronological] Sử dụng khung quá khứ (Ví dụ: 2015-2021) làm bộ gốc, chừa lại khúc đuôi hiện tại (2022-2023) làm bài thi nhằm kiểm tra sức chống chịu lạm phát thực tiễn.",
            "• [Blind Test cách ly] Cắt đứt hoàn toàn thuật toán với tập kiểm thử ngoại suy (Out-of-Sample) để mô phỏng chính xác áp lực rủi ro thật mà Ban Đầu tư sẽ phải đối mặt."
        ]
    },
    {
        "num": "03",
        "title": "SCAN LỖI LẤY LÒNG MÔ HÌNH (ANTI-PERSISTENCE BIAS)",
        "items": [
            "• [Bẫy ML Kinh điển] Hủy bỏ ngay lừa đảo 'Copy paste': Quét xem AI có lười nhác dự đoán nguyên xi hạng tín nhiệm quý này y hệt quý trước để đạt tỷ lệ đúng (Accuracy) 95% hay không.",
            "• [Ma trận Đánh giá ChgAcc] Không dùng Accuracy. Soi thẳng vào F1-Weighted và Change-Accuracy: Yêu cầu AI phải phán đoán trúng khoảnh khắc doanh nghiệp bắt đầu tụt hạng (Downgrade)."
        ]
    },
    {
        "num": "04",
        "title": "PENALTY AUDIT & KIỂM THỬ KHỦNG HOẢNG (STRESS TESTING)",
        "items": [
            "• [Hệ thống Phạt CORN Loss] Đo khoảng cách sai số hạng. Sai khác nhỏ (AA sang A) bị phạt nhẹ. Phạt theo cấp số nhân nếu mô hình phán đoán nhầm Default (Vỡ nợ) thành Investment Grade (An Toàn).",
            "• [Tiêm nhiễu vĩ mô] Bơm tập dữ liệu cực đoan từ SMOTE / TimeGAN vào tầng Test để thử tải bộ phuộc giảm xóc Transformer Bi-LSTM."
        ]
    },
    {
        "num": "05",
        "title": "GIẢI MÃ SHAP AUDIT (HỘP ĐEN EXPLAINABILITY)",
        "items": [
            "• [Truy vết Quyết định Lý tính] Giải mã lý do Hạ Hạng. Mô hình phải chứng minh do sụt biên ròng, suy kiệt Dòng tiền (FCF), thay vì ăn may lấy lý do từ mã Ngành.",
            "• [Ủy ban Thẩm định Nhân tạo] Human-in-the-loop: Chạy quy trình nghiệm thu ngược, báo cáo các biến tài chính bị đổi cực trọng số lên Giám đốc Rủi ro kiểm duyệt."
        ]
    }
]

# Write Main Title
ws.merge_cells("B2:G3")
title_cell = ws.cell(row=2, column=2, value="QUY TRÌNH KIỂM THỬ & NGHIỆM THU MÔ HÌNH TÍN DỤNG\n(MODEL VALIDATION & STRESS TESTING PROTOCOL)")
title_cell.fill = FILL_MAIN
title_cell.font = FONT_MAIN_TITLE
title_cell.alignment = ALIGN_CENTER

curr_row = 5

for i, block in enumerate(flow_data):
    num_items = len(block["items"])
    total_rows = num_items + 1 
    
    end_row = curr_row + total_rows - 1
    
    # Write Step Number
    ws.merge_cells(start_row=curr_row, start_column=2, end_row=end_row, end_column=2)
    num_cell = ws.cell(row=curr_row, column=2, value=block["num"])
    num_cell.fill = FILL_STEP_BG
    num_cell.font = FONT_STEP_NUM
    num_cell.alignment = ALIGN_CENTER
    
    # Write Title
    ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=7)
    th_cell = ws.cell(row=curr_row, column=3, value="  " + block["title"])
    th_cell.fill = FILL_HEADER
    th_cell.font = FONT_STEP_TITLE
    th_cell.alignment = ALIGN_LEFT_NO_INDENT
    for col in range(3, 8):
        c = ws.cell(row=curr_row, column=col)
        c.border = Border(left=Side(style='medium', color="6C3483") if col==3 else None,
                          top=THIN_PALE, right=THIN_PALE if col==7 else None)
    ws.row_dimensions[curr_row].height = 25
    
    # Write Items
    item_row = curr_row + 1
    for item in block["items"]:
        ws.merge_cells(start_row=item_row, start_column=3, end_row=item_row, end_column=7)
        it_cell = ws.cell(row=item_row, column=3, value=item)
        it_cell.fill = FILL_CONTENT
        # Highlight important bracketed text
        if "[" in item and "]" in item:
            it_cell.font = FONT_BODY_BOLD # Basic highlight logic handled dynamically here via color
        else:
            it_cell.font = FONT_BODY
        it_cell.alignment = ALIGN_LEFT
        for col in range(4, 8):
            c = ws.cell(row=item_row, column=col)
            c.border = Border(right=THIN_PALE if col==7 else None, bottom=THIN_PALE)
            
        it_cell.border = Border(left=Side(style='medium', color="6C3483"), bottom=THIN_PALE)
            
        ws.row_dimensions[item_row].height = 32
        item_row += 1
        
    curr_row = end_row + 1
    
    # Draw Arrow if not last
    if i < len(flow_data) - 1:
        ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=7)
        arr_cell = ws.cell(row=curr_row, column=3, value="▼")
        arr_cell.font = FONT_ARROW
        arr_cell.alignment = ALIGN_CENTER
        ws.row_dimensions[curr_row].height = 20
        curr_row += 1
        
# Set column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 22

ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = "6C3483" # Match the purple theme for Testing

wb.save(dest_file)
print(f"Testing Process Protocol added to {dest_file}")
